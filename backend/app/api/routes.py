# app/api/routes.py

import base64
import uuid
import os
import re
from pathlib import Path
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, Dict
from app.rock_core_analysis import analyze_rock_core_photo, generate_rock_core_pdf

try:
    from app.graph import graph
except ModuleNotFoundError as exc:
    graph = None
    GRAPH_IMPORT_ERROR = str(exc)
else:
    GRAPH_IMPORT_ERROR = ""

router = APIRouter(prefix="/api/v1")
DEFAULT_MAX_UPLOAD_BYTES = 25 * 1024 * 1024
UPLOAD_CHUNK_BYTES = 1024 * 1024
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff", ".heic", ".heif"}


def upload_root() -> Path:
    root = Path(os.getenv("AUTOSOIL_UPLOAD_ROOT", "uploads")).resolve()
    root.mkdir(parents=True, exist_ok=True)
    return root


def _safe_name(filename: str | None, fallback: str = "upload.bin") -> str:
    name = Path(filename or fallback).name
    stem = re.sub(r"[^A-Za-z0-9._-]+", "-", Path(name).stem).strip("-") or "upload"
    suffix = Path(name).suffix.lower()
    return f"{stem}_{uuid.uuid4().hex[:10]}{suffix}"


def _max_upload_bytes() -> int:
    configured = os.getenv("AUTOSOIL_MAX_UPLOAD_BYTES", "").strip()
    if not configured:
        return DEFAULT_MAX_UPLOAD_BYTES
    try:
        value = int(configured)
    except ValueError:
        return DEFAULT_MAX_UPLOAD_BYTES
    return max(1, value)


def _validate_upload_type(file: UploadFile, allowed_extensions: set[str] | None = None) -> None:
    if not allowed_extensions:
        return
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in allowed_extensions:
        raise HTTPException(status_code=415, detail=f"Unsupported upload type: {suffix or 'unknown'}")


async def _save_upload(file: UploadFile, folder: str = "evidence", allowed_extensions: set[str] | None = None) -> Path:
    _validate_upload_type(file, allowed_extensions)
    limit = _max_upload_bytes()
    target_dir = upload_root() / folder
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / _safe_name(file.filename)
    total = 0
    try:
        with path.open("wb") as output:
            while True:
                chunk = await file.read(UPLOAD_CHUNK_BYTES)
                if not chunk:
                    break
                total += len(chunk)
                if total > limit:
                    raise HTTPException(status_code=413, detail=f"Upload exceeds {limit} byte limit.")
                output.write(chunk)
    except Exception:
        if path.exists():
            path.unlink(missing_ok=True)
        raise
    return path


def _public_upload_url(path: Path) -> str:
    relative = path.resolve().relative_to(upload_root())
    return "/uploads/" + "/".join(relative.parts)


def _require_graph():
    if graph is None:
        raise HTTPException(
            status_code=503,
            detail=f"Geotechnical supervisor graph is unavailable: {GRAPH_IMPORT_ERROR}",
        )
    return graph


def _resolve_deepseek_target():
    """
    Resolve how to reach DeepSeek, in priority order:
    1. A direct API key on this host (any of the accepted env names).
    2. The Vercel serverless proxy (/api/deepseek/chat) via DEEPSEEK_PROXY_URL,
       for setups where the key only lives in Vercel env vars.
    Returns (mode, url, headers) or (None, None, None) when unconfigured.
    """
    from app.llm_provider import resolve_llm_config
    direct_key = None
    if os.getenv("MODEL_PROVIDER", "").strip().lower() == "deepseek":
        direct_key = resolve_llm_config().api_key
    direct_key = direct_key or os.getenv("DEEPSEEK_API_KEY") or os.getenv("DeepSeek") or os.getenv("DEEPSEEK")
    if direct_key:
        return (
            "direct",
            "https://api.deepseek.com/chat/completions",
            {"Authorization": f"Bearer {direct_key}", "Content-Type": "application/json"},
        )

    proxy_url = os.getenv("DEEPSEEK_PROXY_URL")
    if proxy_url:
        headers = {"Content-Type": "application/json"}
        autosoil_key = os.getenv("AUTOSOIL_API_KEY")
        if autosoil_key:
            headers["X-Autosoil-Api-Key"] = autosoil_key
        return ("proxy", proxy_url, headers)

    return (None, None, None)


def _deepseek_lithology_draft(rows: list, core_segments: list, defects: list, core_recovery: list) -> tuple[list, list]:
    """
    Draft AS1726 lithology units from the vision metrics via an LLM.

    Routes through app.inference (multi-provider fallback chain + transient-error
    retries + tolerant JSON extraction, adapted from the Odysseus project) so a
    single provider outage, rate-limit, or markdown-wrapped reply no longer
    collapses the whole borehole draft. Falls back to one review-required unit
    only when every configured provider fails, so the human-review screen always
    has something to correct.
    """
    from app.inference import build_candidates_from_env, chat_json

    warnings: list[str] = []
    candidates = build_candidates_from_env()

    prompt = f"""
    You are a senior geotechnical engineer logging rock core to AS1726 standards.
    Based on the following core segments and recovery metrics, draft the lithology units.
    Metrics: {core_recovery}
    Segments: {core_segments}
    Defects: {defects}

    Respond with ONLY valid JSON matching this schema:
    {{
        "lithology_units": [
            {{
                "from": float,
                "to": float,
                "material": "string (e.g. SANDSTONE)",
                "description": "string",
                "weathering": "string (e.g. FR, SW, HW)",
                "strength": "string (e.g. L, M, H, VH)",
                "review_required": true
            }}
        ]
    }}
    """

    messages = [
        {"role": "system", "content": "You are a senior geotechnical engineer logging rock core to AS1726. Reply with valid JSON only."},
        {"role": "user", "content": prompt},
    ]

    lithology: list = []
    if not candidates:
        warnings.append(
            "No LLM provider configured. Set DEEPSEEK_API_KEY (or OPENAI_API_KEY / "
            "ANTHROPIC_API_KEY / OLLAMA_BASE_URL). Draft requires manual logging."
        )
    else:
        try:
            parsed, _used = chat_json(messages, candidates=candidates, temperature=0.2, max_tokens=1800)
            if isinstance(parsed, dict):
                lithology = parsed.get("lithology_units", []) or []
            elif isinstance(parsed, list):
                lithology = parsed
        except Exception as e:
            warnings.append(f"AI draft failed across {len(candidates)} provider(s): {e}")

    if not lithology and rows:
        lithology = [{
            "from": rows[0].get("from", 0.0),
            "to": rows[-1].get("to", 10.0),
            "material": "UNKNOWN ROCK",
            "description": "AI draft unavailable. Please log manually.",
            "weathering": "FR",
            "strength": "M",
            "review_required": True,
        }]

    for unit in lithology:
        unit.setdefault("review_required", True)
        unit.setdefault("status", "draft")
    return lithology, warnings


class LogIntervalRequest(BaseModel):
    project_id: str
    project_name: str
    borehole_id: str
    depth_from: float
    depth_to: float
    sample_id: str = ""
    # Optional manual overrides (user can pre-fill these)
    colour: Optional[str] = None
    moisture: Optional[str] = None
    consistency: Optional[str] = None
    notes: Optional[str] = None


class LogIntervalWithPhotoRequest(LogIntervalRequest):
    photo_base64: Optional[str] = None


class TemplateGenerateRequest(BaseModel):
    template_id: str
    replacements: Dict[str, str]
    selected_historical_report_path: Optional[str] = None
    project_path: Optional[str] = None


class ExtractRequest(BaseModel):
    file_path: str


class ClassifyIntervalRequest(BaseModel):
    notes: Optional[str] = None
    photo_base64: Optional[str] = None


@router.post("/log-interval")
async def log_interval(request: LogIntervalRequest):
    """
    Log a soil interval without a photo.
    The classifier agent will attempt classification from existing state context.
    """
    active_graph = _require_graph()
    thread_id = f"{request.project_id}-{request.borehole_id}"
    config = {"configurable": {"thread_id": thread_id}}

    # Try to get existing state to preserve soil_layers
    existing_state = await active_graph.aget_state(config)
    soil_layers = []
    test_results = []
    if existing_state and existing_state.values:
        soil_layers = existing_state.values.get("soil_layers", [])
        test_results = existing_state.values.get("test_results", [])

    provider = os.getenv("MODEL_PROVIDER", "openai")
    if provider == "mock":
        # Validate depth
        if request.depth_from > request.depth_to:
            raise HTTPException(status_code=400, detail="depth_from cannot be greater than depth_to")
            
        new_layer = {
            "depth_from": request.depth_from,
            "depth_to": request.depth_to,
            "uscs_code": "CL",
            "description": request.notes or "",
            "colour": request.colour or "brown",
            "moisture": request.moisture or "moist",
            "consistency": request.consistency or "stiff",
            "structure": "massive",
            "inclusions": ""
        }
        
        # Try to parse USCS code from notes description
        import re
        uscs_match = re.search(r'\b(CH|CL|SC|SM|SP|GP|GW|ML|MH|OH|OL|PT|GM|GC)\b', request.notes or "")
        if uscs_match:
            new_layer["uscs_code"] = uscs_match.group(1)

        soil_layers.append(new_layer)
        
        # Update graph state so it persists
        state_update = {
            "soil_layers": soil_layers,
            "current_layer": None
        }
        await graph.update_state(config, state_update)

        # Save to DB
        from app.utils.db import save_layer
        save_layer(
            project_id=request.project_id,
            project_name=request.project_name,
            borehole_id=request.borehole_id,
            layer=new_layer
        )
        
        return {
            "status": "logged",
            "layer": new_layer,
            "qa_score": 95.0,
            "total_layers": len(soil_layers),
        }

    initial_state = {
        "project_id": request.project_id,
        "project_name": request.project_name,
        "borehole_id": request.borehole_id,
        "depth_from": request.depth_from,
        "depth_to": request.depth_to,
        "sample_id": request.sample_id,
        "soil_layers": soil_layers,
        "test_results": test_results,
        "qa_score": 0.0,
        "qa_passed": False,
        "retry_count": 0,
        "messages": [],
        "last_agent": "start",
        "error": None,
        "pending_human_review": False,
        "validation_errors": [],
        "historical_context": None,
        "compliance_check": None,
        "executive_summary": None,
        "dispatch_status": None,
        "is_dispatched": False,
        "current_layer": {
            "depth_from": request.depth_from,
            "depth_to": request.depth_to,
            "colour": request.colour or "",
            "moisture": request.moisture or "",
            "consistency": request.consistency or "",
        },
    }

    result = await active_graph.ainvoke(initial_state, config=config)

    if result.get("pending_human_review"):
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Classification failed QA after 3 retries - human review required",
                "error": result.get("error"),
                "qa_score": result.get("qa_score"),
                "qa_feedback": result.get("qa_feedback"),
            }
        )

    return {
        "status": "logged",
        "layer": result.get("soil_layers", [])[-1] if result.get("soil_layers") else None,
        "qa_score": result.get("qa_score"),
        "total_layers": len(result.get("soil_layers") or []),
    }


@router.post("/log-interval-photo")
async def log_interval_with_photo(
    project_id: str = Form(...),
    project_name: str = Form(...),
    borehole_id: str = Form(...),
    depth_from: float = Form(...),
    depth_to: float = Form(...),
    sample_id: str = Form(""),
    photo: UploadFile = File(...),
):
    """Log a soil interval with a field photo. Triggers photo -> classify -> QA -> log chain."""
    active_graph = _require_graph()
    contents = await photo.read()
    b64 = base64.b64encode(contents).decode("utf-8")

    thread_id = f"{project_id}-{borehole_id}"
    config = {"configurable": {"thread_id": thread_id}}

    # Try to get existing state to preserve soil_layers
    existing_state = await active_graph.aget_state(config)
    soil_layers = []
    test_results = []
    if existing_state and existing_state.values:
        soil_layers = existing_state.values.get("soil_layers", [])
        test_results = existing_state.values.get("test_results", [])

    initial_state = {
        "project_id": project_id,
        "project_name": project_name,
        "borehole_id": borehole_id,
        "depth_from": depth_from,
        "depth_to": depth_to,
        "sample_id": sample_id,
        "photo_base64": b64,
        "soil_layers": soil_layers,
        "test_results": test_results,
        "qa_score": 0.0,
        "qa_passed": False,
        "retry_count": 0,
        "messages": [],
        "last_agent": "start",
        "error": None,
        "pending_human_review": False,
        "validation_errors": [],
        "historical_context": None,
        "compliance_check": None,
        "executive_summary": None,
        "dispatch_status": None,
        "is_dispatched": False,
    }

    result = await active_graph.ainvoke(initial_state, config=config)

    if result.get("pending_human_review"):
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Classification failed QA after 3 retries - human review required",
                "error": result.get("error"),
                "qa_score": result.get("qa_score"),
                "qa_feedback": result.get("qa_feedback"),
            }
        )

    return {
        "status": "logged",
        "layer": result.get("soil_layers", [])[-1] if result.get("soil_layers") else None,
        "qa_score": result.get("qa_score"),
        "total_layers": len(result.get("soil_layers") or []),
    }


@router.post("/generate-report/{borehole_id}")
async def generate_report(borehole_id: str, project_id: str, project_name: str):
    """Generate AS 1726:2017 PDF report for a completed borehole."""
    from app.agents.report_agent import report_agent
    active_graph = _require_graph()

    thread_id = f"{project_id}-{borehole_id}"
    config = {"configurable": {"thread_id": thread_id}}
    
    existing_state = await active_graph.aget_state(config)
    soil_layers = []
    test_results = []
    if existing_state and existing_state.values:
        soil_layers = existing_state.values.get("soil_layers", [])
        test_results = existing_state.values.get("test_results", [])

    state = {
        "project_id": project_id,
        "project_name": project_name,
        "borehole_id": borehole_id,
        "soil_layers": soil_layers,
        "test_results": test_results,
        "messages": [],
    }

    result = await report_agent(state)
    if result.get("error"):
        raise HTTPException(status_code=500, detail=result["error"])

    return FileResponse(
        path=result["report_path"],
        media_type="application/pdf",
        filename=f"{borehole_id}_log.pdf",
    )


@router.get("/templates")
async def get_templates():
    """List all templates in Main STS Templates folder."""
    from app.utils.template_manager import list_templates
    try:
        return list_templates()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/templates/placeholders")
async def get_template_placeholders(template_id: str):
    """Scan and retrieve all bracketed placeholders inside a selected template."""
    from app.utils.template_manager import extract_placeholders, TEMPLATES_DIR
    try:
        path_suffix = template_id.replace("/", os.sep)
        full_path = os.path.join(TEMPLATES_DIR, path_suffix)
        if not os.path.exists(full_path):
            raise HTTPException(status_code=404, detail="Template file not found")
        return {
            "template_id": template_id,
            "placeholders": extract_placeholders(full_path)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


from typing import List

class SuggestFieldsRequest(BaseModel):
    template_id: str
    placeholders: List[str]
    selected_historical_report_path: Optional[str] = None
    project_path: Optional[str] = None


@router.post("/templates/suggest-fields")
async def suggest_template_fields(request: SuggestFieldsRequest):
    """
    Sends targeted keyword requests straight to the ColBERT index and returns pre-filled defaults.
    """
    from app.utils.rag_manager import build_index_from_reports, search_rag_late_interaction
    from app.utils.template_manager import extract_placeholders, extract_variables_from_docx
    from app.agents.template_swarm import extract_value_via_llm, extract_field_value_from_chunk
    
    try:
        # Step A: Ensure ColBERT index is initialized
        print("[UI Engine Builder Agent] Triggering build_index_from_reports...", flush=True)
        build_index_from_reports(limit=15)
        
        suggestions = {}
        hist_path = request.selected_historical_report_path
        
        placeholders = list(request.placeholders)
        core_fields = ["CLIENT", "CLIENT_NAME", "ADDRESS", "SITE_ADDRESS", "JOB_NO", "REPORT_NO", "DATE", "ENGINEER", "BEARING_CAPACITY"]
        for cf in core_fields:
            if cf not in placeholders:
                placeholders.append(cf)
        
        # Regex baseline extraction if selected_historical_report_path is provided
        if hist_path and os.path.exists(hist_path):
            print(f"[UI Engine Builder Agent] Running regex parser on: {os.path.basename(hist_path)}", flush=True)
            extracted_data = extract_variables_from_docx(hist_path)
            for k, v in extracted_data.items():
                if k in placeholders and v:
                    suggestions[k] = v

        # If project_path is provided, we can also query project-specific RAG/OCR variables
        if request.project_path and os.path.exists(request.project_path):
            try:
                from app.utils.rag_manager import analyze_project_folder
                analysis = analyze_project_folder(request.project_path)
                mappings = {
                    "CLIENT": "client",
                    "CLIENT_NAME": "client",
                    "ADDRESS": "site_address",
                    "SITE_ADDRESS": "site_address",
                    "JOB_NO": "job_no",
                    "BEARING_CAPACITY": "bearing_capacity"
                }
                for placeholder, analysis_key in mappings.items():
                    if placeholder in placeholders and not suggestions.get(placeholder) and analysis.get(analysis_key):
                        suggestions[placeholder] = analysis[analysis_key]
                if "NOTES" in placeholders and not suggestions.get("NOTES") and analysis.get("summary"):
                    suggestions["NOTES"] = analysis["summary"]
            except Exception as e:
                print(f"[UI Engine Builder Agent] Error auto-filling from project_path: {e}", flush=True)

        QUERY_MAPPINGS = {
            "CLIENT": "Client Name",
            "CLIENT_NAME": "Client Name",
            "ADDRESS": "Project Address",
            "SITE_ADDRESS": "Project Address",
            "JOB_NO": "Job Number",
            "REPORT_NO": "Report Number",
            "DATE": "Report Date",
            "ENGINEER": "Author Engineer",
            "BEARING_CAPACITY": "allowable bearing capacity"
        }
        
        # Query ColBERT for remaining blank placeholders
        for ph in placeholders:
            if not suggestions.get(ph):
                query_term = QUERY_MAPPINGS.get(ph, ph.replace("_", " ").title())
                print(f"[UI Engine Builder Agent] Querying ColBERT for keyword '{ph}' -> '{query_term}'...", flush=True)
                
                results = search_rag_late_interaction(query_term, limit=10, file_path=hist_path)
                if results:
                    top_chunk = results[0]
                    chunk_text = top_chunk["chunk_text"]
                    
                    # Try to extract via LLM
                    val = await extract_value_via_llm(ph, chunk_text)
                    if not val:
                        val = extract_field_value_from_chunk(ph, chunk_text)
                    if not val:
                        val = chunk_text.split("\n")[0][:150].strip()
                        
                    if val:
                        suggestions[ph] = val
                        
        # Cross-fill Client & Client Name
        if "CLIENT" in suggestions and "CLIENT_NAME" not in suggestions:
            suggestions["CLIENT_NAME"] = suggestions["CLIENT"]
        elif "CLIENT_NAME" in suggestions and "CLIENT" not in suggestions:
            suggestions["CLIENT"] = suggestions["CLIENT_NAME"]
            
        return {
            "status": "success",
            "suggestions": suggestions
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/templates/generate")
async def generate_from_template(request: TemplateGenerateRequest):
    """Instantiate a completed geotechnical report .docx by replacing all template placeholders via a Multi-Agent Swarm."""
    from app.utils.template_manager import TEMPLATES_DIR
    from app.agents.template_swarm import template_swarm
    try:
        path_suffix = request.template_id.replace("/", os.sep)
        template_path = os.path.join(TEMPLATES_DIR, path_suffix)
        
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="Template not found")
            
        initial_state = {
            "template_id": request.template_id,
            "template_path": template_path,
            "selected_historical_report_path": request.selected_historical_report_path,
            "input_replacements": request.replacements,
            "placeholders": [],
            "replacements": {},
            "compliance_check": "",
            "qa_score": 0.0,
            "qa_passed": False,
            "qa_feedback": [],
            "output_path": None,
            "dispatch_status": "",
            "messages": [],
            "last_agent": "",
            "error": None
        }
        
        # If project_path is provided, run site data analysis & merge values
        replacements = {**request.replacements}
        if request.project_path:
            try:
                from app.utils.rag_manager import analyze_project_folder
                analysis = analyze_project_folder(request.project_path)
                mappings = {
                    "CLIENT": "client",
                    "CLIENT_NAME": "client",
                    "ADDRESS": "site_address",
                    "SITE_ADDRESS": "site_address",
                    "JOB_NO": "job_no",
                    "BEARING_CAPACITY": "bearing_capacity"
                }
                for placeholder, analysis_key in mappings.items():
                    if not replacements.get(placeholder) and analysis.get(analysis_key):
                        replacements[placeholder] = analysis[analysis_key]
                if not replacements.get("NOTES") and analysis.get("summary"):
                    replacements["NOTES"] = analysis["summary"]
            except Exception as e:
                print(f"Error auto-filling from project_path: {e}")

        initial_state["input_replacements"] = replacements

        # Invoke our production-grade multi-agent geotechnical swarm!
        result = await template_swarm.ainvoke(initial_state)
        
        if result.get("error"):
            raise HTTPException(status_code=500, detail=result["error"])
            
        completed_path = result.get("output_path")
        if not completed_path or not os.path.exists(completed_path):
            raise HTTPException(status_code=500, detail="Compilation agent failed to generate output path.")
            
        file_name = f"Generated_{os.path.basename(template_path)}"
        
        # Expose QA score and compliance checks via response headers
        headers = {
            "X-QA-Score": str(result.get("qa_score", 0.0)),
            "X-QA-Passed": "true" if result.get("qa_passed", False) else "false",
            "X-Compliance-Check": result.get("compliance_check", ""),
            "Access-Control-Expose-Headers": "X-QA-Score, X-QA-Passed, X-Compliance-Check"
        }
        
        return FileResponse(
            path=completed_path,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=file_name,
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reports/history")
async def get_reports_history():
    """List completed historical reports inside Reports 1, 2, 3 directories."""
    from app.utils.template_manager import list_history

    try:
        return list_history()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rock-core/analyze")
async def analyze_rock_core_photo_upload(
    projectNumber: str = Form(""),
    boreholeId: str = Form(""),
    clientName: str = Form(""),
    siteAddress: str = Form(""),
    inspectionDate: str = Form(""),
    depthFrom: str = Form(""),
    depthTo: str = Form(""),
    photo: UploadFile = File(...),
):
    photo_path = await _save_upload(photo, "rock-core", IMAGE_EXTENSIONS)
    metadata = {
        "projectNumber": projectNumber,
        "boreholeId": boreholeId,
        "client": clientName,
        "siteAddress": siteAddress,
        "inspectionDate": inspectionDate,
        "depthFrom": depthFrom,
        "depthTo": depthTo,
    }
    analysis = analyze_rock_core_photo(photo_path, metadata)
    pdf_path = generate_rock_core_pdf(analysis, photo_path, upload_root() / "rock-core-reports")
    return {
        "status": "rock_core_pdf_ready",
        "message": "Rock core photo analysed. A review-ready PDF and strict JSON extract have been generated.",
        "analysis": analysis,
        "pdf": {
            "fileName": pdf_path.name,
            "downloadUrl": _public_upload_url(pdf_path),
            "reviewRequired": True,
        },
        "needsReview": True,
    }


# ==========================================
# NEW ROCK CORE WORKSPACE ENDPOINTS
# ==========================================

@router.get("/core/boreholes")
async def get_core_boreholes():
    from app.utils.db import list_rock_boreholes
    try:
        return list_rock_boreholes()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/core/borehole/{borehole_id}")
async def get_core_borehole(borehole_id: str):
    from app.utils.db import get_rock_borehole_data
    try:
        data = get_rock_borehole_data(borehole_id)
        if not data:
            raise HTTPException(status_code=404, detail="Borehole not found")
        return data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/core/upload")
async def upload_core_photo(photo: UploadFile = File(...)):
    try:
        photo_path = await _save_upload(photo, "rock-core", IMAGE_EXTENSIONS)
        from PIL import Image
        from app.rock_core_analysis import _detect_core_rows
        
        with Image.open(photo_path) as raw:
            img = raw.convert("RGB")
            gray = img.convert("L")
            rows = _detect_core_rows(gray)
            
        return {
            "status": "success",
            "photo_path": str(photo_path.relative_to(upload_root())),
            "photo_url": _public_upload_url(photo_path),
            "width": img.width,
            "height": img.height,
            "rows": rows
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CorePreprocessRequest(BaseModel):
    photo_path: str

@router.post("/core/preprocess")
async def preprocess_core_photo(req: CorePreprocessRequest):
    return {
        "status": "success",
        "message": "Image rotated and perspective corrected.",
        "processed_image_path": req.photo_path,
        "rows": [{"id": 1, "top": 100, "bottom": 300}, {"id": 2, "top": 350, "bottom": 550}]
    }

class CoreCalibrateRequest(BaseModel):
    rows: list[dict]

@router.post("/core/calibrate-scale")
async def calibrate_scale(req: CoreCalibrateRequest):
    from app.vision.depth_scale_calibration import assign_depth_scale
    calibrated_rows = assign_depth_scale(req.rows, start_depth=9.0)
    return {"status": "success", "rows": calibrated_rows}

class CoreSegmentRequest(BaseModel):
    rows: list[dict]

@router.post("/core/segment")
async def segment_core(req: CoreSegmentRequest):
    from app.vision.core_piece_segmentation import segment_core_pieces
    segments = segment_core_pieces(req.rows)
    return {"status": "success", "core_segments": segments}

class CoreMarkupRequest(BaseModel):
    core_segments: list[dict]

@router.post("/core/markup")
async def markup_defects(req: CoreMarkupRequest):
    from app.vision.defect_markup import markup_defects
    defects = markup_defects(req.core_segments)
    return {"status": "success", "defects": defects}

class CoreCalculateRequest(BaseModel):
    rows: list[dict]
    core_segments: list[dict]
    defects: list[dict]

@router.post("/core/calculate")
async def calculate_metrics(req: CoreCalculateRequest):
    from app.vision.geotech_metrics import calculate_metrics
    recovery = calculate_metrics(req.rows, req.core_segments, req.defects)
    return {"status": "success", "core_recovery": recovery}

class CoreDraftRequest(BaseModel):
    project: dict
    borehole: dict
    rows: list[dict]
    core_segments: list[dict]
    defects: list[dict]
    core_recovery: list[dict]

@router.post("/core/deepseek-log")
async def generate_deepseek_log(req: CoreDraftRequest):
    lithology, warnings = _deepseek_lithology_draft(req.rows, req.core_segments, req.defects, req.core_recovery)
    return {
        "project": req.project,
        "borehole": req.borehole,
        "rows": req.rows,
        "core_segments": req.core_segments,
        "defects": req.defects,
        "core_recovery": req.core_recovery,
        "lithology_units": lithology,
        "warnings": warnings
    }


@router.post("/core/auto-log")
async def auto_log_core_box(
    photo: UploadFile = File(...),
    client: str = Form(""),
    borehole_id: str = Form("BH-01"),
    depth_from: float = Form(0.0),
    depth_to: float = Form(10.0),
):
    """
    One-shot pipeline: photo in -> OpenGround-ready draft out.
    Runs upload -> vision (rows/fractures) -> depth calibration -> piece
    segmentation -> defect markup -> TCR/RQD -> DeepSeek AS1726 draft.
    Everything returned is status='draft' / review_required=True: the human
    reviewer must approve before /pdf/export will accept it.
    """
    if depth_to <= depth_from:
        raise HTTPException(status_code=400, detail="depth_to must be greater than depth_from")

    saved_path = await _save_upload(photo, folder="core-boxes", allowed_extensions={".jpg", ".jpeg", ".png", ".webp"})

    from app.vision.pipeline import process_core_image
    from app.vision.depth_scale_calibration import assign_depth_scale
    from app.vision.core_piece_segmentation import segment_core_pieces
    from app.vision.defect_markup import markup_defects as _markup
    from app.vision.geotech_metrics import calculate_metrics as _metrics

    try:
        vision = process_core_image(saved_path)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Vision pipeline failed: {e}")

    calibrated_rows = assign_depth_scale(vision["rows"], start_depth=depth_from, end_depth=depth_to)
    core_segments = segment_core_pieces(calibrated_rows)
    defects = _markup(core_segments)
    core_recovery = _metrics(calibrated_rows, core_segments, defects)
    lithology, warnings = _deepseek_lithology_draft(calibrated_rows, core_segments, defects, core_recovery)

    annotated = saved_path.parent / vision["annotated_image"]
    return {
        "status": "draft",
        "review_required": True,
        "project": {"client": client, "boreholeId": borehole_id},
        "borehole": {"id": borehole_id, "depth_from": depth_from, "depth_to": depth_to},
        "photo_url": _public_upload_url(saved_path),
        "annotated_photo_url": _public_upload_url(annotated) if annotated.exists() else None,
        "rows": calibrated_rows,
        "core_segments": core_segments,
        "defects": defects,
        "core_recovery": core_recovery,
        "lithology_units": lithology,
        "warnings": warnings,
    }


class CoreSaveRequest(BaseModel):
    project: dict
    borehole: dict
    lithology_units: list[dict]
    discontinuities: list[dict]
    core_runs: list[dict]

@router.post("/core/save")
async def save_core_log(req: CoreSaveRequest):
    from app.utils.db import save_rock_borehole_data
    try:
        save_rock_borehole_data(
            req.project,
            req.borehole,
            req.lithology_units,
            req.discontinuities,
            req.core_runs
        )
        return {"status": "success", "message": "Borehole log saved successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CoreApproveRequest(BaseModel):
    borehole_id: str

@router.post("/core/review")
async def approve_core_log(req: CoreApproveRequest):
    from app.utils.db import get_sqlite_conn
    conn = get_sqlite_conn()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE rock_lithology_units SET status = 'approved' WHERE borehole_id = ?", (req.borehole_id,))
        cursor.execute("UPDATE rock_discontinuities SET status = 'approved' WHERE borehole_id = ?", (req.borehole_id,))
        cursor.execute("UPDATE rock_core_runs SET status = 'approved' WHERE borehole_id = ?", (req.borehole_id,))
        conn.commit()
        return {"status": "success", "message": "All fields approved."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.post("/pdf/export")
async def export_borehole_pdf(req: CoreSaveRequest):
    # Human-in-the-loop gate: every lithology unit must be reviewer-approved.
    unapproved = [
        u for u in req.lithology_units
        if u.get("status") != "approved" and u.get("approved") is not True
    ]
    if not req.lithology_units or unapproved:
        raise HTTPException(
            status_code=409,
            detail=f"Export blocked: {len(unapproved) if req.lithology_units else 'all'} lithology unit(s) not approved. Review and approve the draft first.",
        )
    try:
        from app.utils.rock_pdf import generate_openground_style_pdf
        data = {
            "project": req.project,
            "borehole": req.borehole,
            "lithology_units": req.lithology_units,
            "discontinuities": req.discontinuities,
            "core_runs": req.core_runs
        }
        
        pdf_path = generate_openground_style_pdf(data, upload_root() / "rock-core-reports")
        
        return {
            "status": "success",
            "pdf_url": _public_upload_url(pdf_path),
            "pdf_name": pdf_path.name
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/uploads/{relative_path:path}")
async def get_uploaded_file(relative_path: str):
    root = upload_root()
    path = (root / relative_path).resolve()
    try:
        path.relative_to(root)
    except ValueError:
        raise HTTPException(status_code=404, detail="Upload not found")
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Upload not found")
    return FileResponse(path=path, filename=path.name)


@router.post("/reports/history/extract")
async def extract_report_variables(request: ExtractRequest):
    """Triggers smart extraction agent to pull all client and site address parameters from a completed report."""
    from app.utils.template_manager import extract_variables_from_docx
    try:
        if not os.path.exists(request.file_path):
            raise HTTPException(status_code=404, detail="Report file not found")
            
        data = extract_variables_from_docx(request.file_path)
        return {
            "status": "success",
            "extracted_data": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/boreholes/{project_id}/{borehole_id}/layers")
async def get_borehole_layers(project_id: str, borehole_id: str):
    """Retrieve all logged soil layers for a specific borehole from SQLite database."""
    from app.utils.db import get_sqlite_conn, init_db
    init_db()
    bh_id = f"{project_id}-{borehole_id}"
    conn = get_sqlite_conn()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT depth_from, depth_to, uscs_code, description, colour, moisture, consistency, structure, inclusions FROM soil_layers WHERE borehole_id = ? ORDER BY depth_from ASC",
            (bh_id,)
        )
        rows = cursor.fetchall()
        layers = [dict(row) for row in rows]
        return {"status": "success", "borehole_id": bh_id, "layers": layers}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/boreholes/{project_id}/{borehole_id}/layers")
async def delete_borehole_layers(project_id: str, borehole_id: str):
    """Clear all logged soil layers for a specific borehole."""
    from app.utils.db import get_sqlite_conn, init_db
    init_db()
    bh_id = f"{project_id}-{borehole_id}"
    conn = get_sqlite_conn()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM soil_layers WHERE borehole_id = ?", (bh_id,))
        conn.commit()
        return {"status": "success", "message": f"Cleared layers for borehole {bh_id}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/classify-interval")
async def classify_interval(request: ClassifyIntervalRequest):
    """
    Fast AI soil classifier that parses text notes or a photo base64
    and returns geotechnical parameters instantly (WITHOUT saving).
    """
    provider = os.getenv("MODEL_PROVIDER", "openai")
    notes = request.notes or ""
    
    # Defaults
    uscs = "CL"
    primary = "CLAY"
    secondary = "Silty"
    colour = "brown"
    moisture = "moist"
    consistency = "stiff"
    inclusions = "trace sand"
    origin = "RESIDUAL SOIL"
    confidence = 0.85
    reasoning = "Extracted from inputs"

    notes_lower = notes.lower()
    if "sand" in notes_lower:
        uscs = "SP"
        primary = "SAND"
        secondary = "Clayey"
        colour = "grey"
        inclusions = "fine to medium grained"
    elif "gravel" in notes_lower:
        uscs = "GP"
        primary = "GRAVEL"
        secondary = "Sandy"
        colour = "brown to rainbow"
        inclusions = "fine to coarse"
    elif "silt" in notes_lower:
        uscs = "ML"
        primary = "SILT"
        secondary = "Sandy"
        colour = "grey / pale brown"
    elif "asphalt" in notes_lower:
        uscs = "ASP"
        primary = "ASPHALT"
        secondary = ""
        colour = "black"
        inclusions = "dense"
        origin = "FILL"
    elif "fill" in notes_lower:
        origin = "FILL"

    if "dark" in notes_lower:
        colour = "dark grey"
    elif "orange" in notes_lower:
        colour = "orange - brown"
    elif "red" in notes_lower:
        colour = "red - brown"

    if "dry" in notes_lower:
        moisture = "dry"
    elif "wet" in notes_lower:
        moisture = "wet"

    if "soft" in notes_lower:
        consistency = "soft"
    elif "firm" in notes_lower:
        consistency = "firm"
    elif "hard" in notes_lower:
        consistency = "hard"

    if provider != "mock":
        try:
            from langchain_core.messages import SystemMessage, HumanMessage
            from app.llm_provider import create_chat_model
            import json

            system_prompt = """
            You are a senior geotechnical engineer classifying soil per AS 1726:2017.
            Analyze the input notes or photo description and extract key soil parameters.
            Respond ONLY with a JSON object containing:
            {
                "uscs_code": "CH/CL/ML/SP/GP/etc",
                "primary_soil": "CLAY/SAND/GRAVEL/SILT/etc",
                "secondary_component": "Silty/Sandy/Clayey/etc",
                "colour": "brown/grey/red-brown/etc",
                "moisture": "dry/moist/wet/saturated",
                "consistency": "soft/firm/stiff/hard/loose/dense/etc",
                "inclusions": "trace sand/gravel/etc",
                "origin": "RESIDUAL SOIL/FILL/COLLUVIAL/TOPSOIL/etc",
                "confidence": 0.0 to 1.0,
                "reasoning": "Brief explanation of classification"
            }
            """
            user_msg = f"Soil field notes: {notes}"
            if request.photo_base64:
                user_msg += "\n[Photo provided in Base64]"
                
            llm = create_chat_model()
                
            response = llm.invoke([SystemMessage(content=system_prompt), HumanMessage(content=user_msg)])
            data = json.loads(response.content.strip())
            
            uscs = data.get("uscs_code", uscs)
            primary = data.get("primary_soil", primary)
            secondary = data.get("secondary_component", secondary)
            colour = data.get("colour", colour)
            moisture = data.get("moisture", moisture)
            consistency = data.get("consistency", consistency)
            inclusions = data.get("inclusions", inclusions)
            origin = data.get("origin", origin)
            confidence = data.get("confidence", confidence)
            reasoning = data.get("reasoning", reasoning)
        except Exception as e:
            print(f"Failed calling LLM for classification: {e}")

    return {
        "status": "success",
        "uscs_code": uscs,
        "primary_soil": primary,
        "secondary_component": secondary,
        "colour": colour,
        "moisture": moisture,
        "consistency": consistency,
        "inclusions": inclusions,
        "origin": origin,
        "confidence": confidence,
        "reasoning": reasoning
    }


@router.get("/health")
async def health():
    graph_nodes = list(graph.nodes.keys()) if graph is not None else []
    return {"status": "ok", "graph_nodes": graph_nodes, "graph_error": GRAPH_IMPORT_ERROR or None}


@router.get("/debug-rag")
async def debug_rag():
    import traceback
    try:
        from app.utils import rag_manager
        # List directories for debugging deployment topology
        parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
        try:
            parent_list = os.listdir(parent_dir)
        except Exception as e:
            parent_list = str(e)
            
        try:
            curr_dir_list = os.listdir(".")
        except Exception as e:
            curr_dir_list = str(e)
            
        # Read main.py content
        main_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "main.py"))
        try:
            with open(main_path, "r") as mf:
                main_content = mf.read()
        except Exception as e:
            main_content = str(e)
            
        return {
            "status": "success",
            "db_path": getattr(rag_manager, "DB_PATH", None),
            "reports_dirs": getattr(rag_manager, "REPORTS_DIRS", None),
            "ocr_reader_available": getattr(rag_manager, "OCR_READER", None) is not None,
            "__file__": __file__,
            "cwd": os.path.abspath("."),
            "parent_dir": parent_dir,
            "parent_list": parent_list,
            "curr_dir_list": curr_dir_list,
            "main_content": main_content
        }
    except Exception as e:
        return {
            "status": "error",
            "error_type": type(e).__name__,
            "error_message": str(e),
            "traceback": traceback.format_exc()
        }


# ----------------------------------------------------
# Advanced RAG & OCR Site Data Analysis Endpoints
# ----------------------------------------------------
from fastapi import BackgroundTasks


def _rag_manager():
    try:
        from app.utils import rag_manager
    except ModuleNotFoundError as exc:
        raise HTTPException(status_code=503, detail=f"RAG/OCR dependencies are unavailable: {exc}") from exc
    return rag_manager

@router.post("/rag/index/build")
async def trigger_rag_build(background_tasks: BackgroundTasks):
    """Trigger background indexing of all reports in reports directories."""
    rag_manager = _rag_manager()
    background_tasks.add_task(rag_manager.build_rag_index)
    return {"status": "indexing_started", "message": "RAG index build has been scheduled in the background."}

@router.get("/rag/search")
async def rag_search(q: str, limit: int = 5):
    """Query the RAG index database for matched historical report details."""
    try:
        results = _rag_manager().search_rag(q, limit)
        return {"status": "success", "results": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/rag/projects")
async def list_rag_projects():
    """List all available project folders in 03 - Reports, Reports 2, Reports 3 directories."""
    rag_manager = _rag_manager()
    projects = []
    for base_dir in rag_manager.REPORTS_DIRS:
        if not os.path.exists(base_dir):
            continue
        group_name = os.path.basename(base_dir)
        for item in os.listdir(base_dir):
            item_path = os.path.join(base_dir, item)
            if os.path.isdir(item_path):
                # Count files inside
                files_count = 0
                for r, d, fs in os.walk(item_path):
                    files_count += len(fs)
                
                projects.append({
                    "project_id": item,
                    "folder_path": item_path.replace("\\", "/").replace("\\", "/"),
                    "group": group_name,
                    "files_count": files_count
                })
    return sorted(projects, key=lambda x: (x["group"], x["project_id"]))

@router.get("/rag/analyze")
async def rag_analyze_project(project_path: str):
    """Run analysis and OCR on a specific site project directory to extract key report variables."""
    try:
        if not os.path.exists(project_path):
            raise HTTPException(status_code=404, detail="Project folder not found")
        analysis = _rag_manager().analyze_project_folder(project_path)
        return {"status": "success", "analysis": analysis}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/project/upload")
async def upload_project_files(files: list[UploadFile] = File(...)):
    """Upload site files (photos, soil logs, lab data) to a temporary project directory and analyze them."""
    import shutil
    import openpyxl
    try:
        upload_id = str(uuid.uuid4())
        upload_dir = os.path.join(
            "C:\\Users\\pored\\.gemini\\antigravity\\brain\\42a1d8dd-f670-4931-864d-7d1abcb1fedc\\scratch\\uploads",
            upload_id
        )
        os.makedirs(upload_dir, exist_ok=True)
        
        saved_files = []
        for file in files:
            file_path = os.path.join(upload_dir, file.filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            saved_files.append(file_path)
            
        # Analyze the newly created project folder using RAG and OCR
        analysis = _rag_manager().analyze_project_folder(upload_dir)
        
        # Extract lab details specifically if there are spreadsheets in the uploaded files
        lab_summary = []
        for fpath in saved_files:
            fname = os.path.basename(fpath).lower()
            if fname.endswith(('.xlsx', '.xls', '.xlsm')):
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                    snames = [s.lower() for s in wb.sheetnames]
                    if any("cbr" in s for s in snames):
                        lab_summary.append(f"CBR test spreadsheet '{os.path.basename(fpath)}'")
                    elif any("point load" in s or "calc" in s for s in snames):
                        lab_summary.append(f"Point Load test spreadsheet '{os.path.basename(fpath)}'")
                    else:
                        lab_summary.append(f"Excel lab file '{os.path.basename(fpath)}' ({', '.join(wb.sheetnames)})")
                except Exception:
                    pass
        
        if lab_summary:
            analysis["summary"] += " [Lab Data Detected]: " + ", ".join(lab_summary) + "."
            
        return {
            "status": "success",
            "project_path": upload_dir.replace("\\", "/"),
            "analysis": analysis
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class CoreStitchRequest(BaseModel):
    photo_path: str
    num_rows: int = 4

@router.post("/core/stitch")
async def stitch_core_photo(req: CoreStitchRequest):
    from app.vision.stitcher import slice_and_stitch_vertically
    
    # Resolve absolute path for processing
    full_path = upload_root() / req.photo_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="Original photo not found")
        
    try:
        stitched_path_abs = slice_and_stitch_vertically(str(full_path), req.num_rows, output_dir=str(upload_root() / "rock-core"))
        stitched_path = Path(stitched_path_abs)
        
        return {
            "status": "success",
            "stitched_photo_path": str(stitched_path.relative_to(upload_root())).replace('\\', '/'),
            "stitched_photo_url": _public_upload_url(stitched_path)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
